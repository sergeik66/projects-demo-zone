import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# Create workbook
wb = Workbook()
ws1 = wb.active
ws1.title = "Dataset Metadata"

# Sheet 1: Dataset Metadata
ws1.append(["dataset_name", "description", "domain", "sub_domain", "data_category", "sensitivity", "retention_period"])
ws1.append(["inspayrol",      "Insurance payroll data",    "Finance",     "Payroll",   "Transactional", "Internal",     "7 years"])
ws1.append(["carrier",        "Carrier information",       "Operations",  "Vendor",    "Reference",     "Public",       "Indefinite"])
ws1.append(["policy_header",  "Policy headers",            "Underwriting","Policy",    "Transactional", "Confidential", "10 years"])

# Header styling
header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
for cell in ws1[1]:
    cell.fill = header_fill
    cell.font = Font(bold=True)

# Helper function (now correct!)
def make_sheet(title, header, data_rows):
    ws = wb.create_sheet(title)
    ws.append(header)
    for row in data_rows:
        ws.append(row)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
    return ws

# Sheets 2–4
make_sheet("Source System Properties",
           ["dataset_name", "source_system", "source_table", "extraction_method", "frequency"],
           [["inspayrol","HRIS","payroll_table","API","Daily"],
            ["carrier","VendorDB","carriers","CSV Export","Weekly"],
            ["policy_header","PolicySys","headers","SQL Query","Real-time"]])

make_sheet("Raw Properties",
           ["dataset_name","raw_schema","raw_columns","raw_volume_estimate","raw_update_frequency"],
           [["inspayrol","payroll_raw","emp_id,salary,date","1M rows","Daily"],
            ["carrier","carrier_raw","carrier_id,name,contact","10K rows","Weekly"],
            ["policy_header","policy_raw","policy_id,type,status","500K rows","Real-time"]])

make_sheet("Curated Properties",
           ["dataset_name","curated_schema","curated_columns","curated_pii_fields","curated_quality_checks"],
           [["inspayrol","payroll_curated","emp_id,net_pay,period","emp_id","Row count match, Null checks"],
            ["carrier","carrier_curated","carrier_id,verified_name","None","Duplicate check"],
            ["policy_header","policy_curated","policy_id,active_flag","policy_id","Freshness SLA: 1h"]])

# Sheet 5: Data Contracts
ws5 = wb.create_sheet("Data Contracts")
ws5.append(["dataset_name","owner_team","lifecycle","freshness_sla","completeness_sla",
            "validity_sla","uniqueness_sla","timeliness_sla","criticality",
            "downstream_consumers","slr_required","data_lineage_notes"])

sample_data = [
    ["inspayrol",     "Finance Team",  "Active", "Daily",      "95%", "99%",   "100%", "24h", "High",     "Reporting, Analytics", "Yes", "HRIS to Raw to Curated"],
    ["carrier",       "Operations",    "Active", "Weekly",     "98%", "100%",  "100%", "7d",  "Medium",   "Vendor Portal",       "No",  "VendorDB to Raw"],
    ["policy_header","Underwriting",  "Active", "Real-time",  "99%", "99.5%", "100%", "1h",  "Critical", "Claims, Billing",     "Yes", "PolicySys to Raw to Curated"]
]
for row in sample_data:
    ws5.append(row)

# Header styling for Data Contracts
for cell in ws5[1]:
    cell.fill = header_fill
    cell.font = Font(bold=True)

# Dropdown validations
dropdowns = [
    ("C2:C1000", '"Active,Deprecated,Archived"'),
    ("D2:D1000", '"Real-time,Hourly,Daily,Weekly,Monthly"'),
    ("E2:E1000", '"90%,95%,98%,99%,100%"'),
    ("F2:G1000", '"95%,98%,99%,99.5%,100%"'),  # validity + uniqueness
    ("H2:H1000", '"1h,6h,24h,7d,30d"'),
    ("I2:I1000", '"Low,Medium,High,Critical"'),
    ("K2:K1000", '"Yes,No"')
]
for range_addr, formula in dropdowns:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.add(range_addr)
    ws5.add_data_validation(dv)

# Conditional formatting: Critical = light red
red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
rule = CellIsRule(operator='equal', formula=['"Critical"'], stopIfTrue=True, fill=red_fill)
ws5.conditional_formatting.add("I2:I1000", rule)

# Auto-size columns
for ws in wb.worksheets:
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

# Save the file
wb.save("source_aligned_master_config.xlsx")
print("SUCCESS! Your Excel file is ready: source_aligned_master_config.xlsx")
